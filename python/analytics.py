import os, re
from datetime import datetime, timedelta
import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine, text
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg2://postgres:admin@localhost:5432/csgo")


os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

plt.rcParams.update({"figure.figsize": (9, 5), "axes.grid": True, "grid.alpha": 0.25, "font.size": 11})

engine = create_engine(DATABASE_URL)

def fetch_df(sql: str, params=None) -> pd.DataFrame:
    with engine.begin() as conn:
        return pd.read_sql(text(sql), conn, params=params or {})

def save_plot(df: pd.DataFrame, fig, filename: str, note: str):
    if df is None or df.empty:
        print(f"[WARN] no data -> skip {filename} | {note}")
        plt.close(fig); return
    path = os.path.join("charts", filename)
    fig.tight_layout(); fig.savefig(path, dpi=160); plt.close(fig)
    print(f"[OK] rows={len(df)} | saved charts/{filename} | {note}")

def save_html(fig, filename: str):
    path = os.path.join("charts", filename)
    fig.write_html(path, include_plotlyjs="cdn")
    print(f"[OK] saved charts/{filename} (HTML)")


EVENT_NAME_CACHE = {}
def get_event_name(event_id: int) -> str:
    if event_id in EVENT_NAME_CACHE: return EVENT_NAME_CACHE[event_id]
    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT event_name FROM events WHERE event_id = :e LIMIT 1
        """), {"e": int(event_id)}).fetchone()
    name = (row[0].strip() if row and row[0] else f"Event {event_id}")
    EVENT_NAME_CACHE[event_id] = name
    return name

def slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")

# --- SQL ---
SQL_PIE = """
SELECT r.map, COUNT(*) AS maps_played
FROM results r
JOIN matches m ON m.match_id = r.match_id
WHERE r.event_id = :event_id
  AND COALESCE(r.map,'') NOT IN ('Default','Unknown','')
GROUP BY r.map
ORDER BY maps_played DESC;
"""
SQL_BAR = """
WITH wins AS (
  SELECT DISTINCT r.match_id,
         CASE WHEN r.match_winner = 1 THEN m.team_1
              WHEN r.match_winner = 2 THEN m.team_2 END AS winner_team
  FROM results r
  JOIN matches m ON m.match_id = r.match_id
  WHERE r.event_id = :event_id
)
SELECT winner_team AS team, COUNT(*) AS wins
FROM wins
WHERE winner_team IS NOT NULL
GROUP BY winner_team
ORDER BY wins DESC
LIMIT 10;
"""
SQL_BARH = """
SELECT pr.player_name, pr.team,
       ROUND(AVG(pr.rating), 2) AS avg_rating,
       COUNT(*) AS maps_played
FROM players_raw pr
JOIN results r ON r.match_id = pr.match_id AND r.event_id = pr.event_id
JOIN matches m ON m.match_id = pr.match_id
WHERE pr.event_id = :event_id
GROUP BY pr.player_name, pr.team
HAVING COUNT(*) >= :min_maps
ORDER BY avg_rating DESC
LIMIT 15;
"""
SQL_LINE = """
SELECT m.match_date::date AS d,
       SUM(
           CASE
             WHEN r.match_winner = 1 AND m.team_1 = :team THEN r.result_1
             WHEN r.match_winner = 2 AND m.team_2 = :team THEN r.result_2
             WHEN m.team_1 = :team THEN r.result_1
             WHEN m.team_2 = :team THEN r.result_2
             ELSE 0
           END
       ) AS rounds_won
FROM results r
JOIN matches m ON m.match_id = r.match_id
WHERE EXTRACT(YEAR FROM m.match_date) = :year
GROUP BY d
ORDER BY d;
"""
SQL_HIST = """
SELECT (r.result_1 + r.result_2) AS total_rounds
FROM results r
JOIN matches m ON m.match_id = r.match_id
WHERE r.event_id = :event_id;
"""
SQL_SCATTER = """
WITH team_rounds AS (
  SELECT r.match_id, m.team_1 AS team, r.result_1 AS rounds_won
  FROM results r
  JOIN matches m ON m.match_id = r.match_id
  WHERE r.event_id = :event_id
  UNION ALL
  SELECT r.match_id, m.team_2 AS team, r.result_2 AS rounds_won
  FROM results r
  JOIN matches m ON m.match_id = r.match_id
  WHERE r.event_id = :event_id
),
best_player AS (
  SELECT pr.match_id, pr.team, MAX(pr.rating) AS best_rating
  FROM players_raw pr
  WHERE pr.event_id = :event_id
  GROUP BY pr.match_id, pr.team
)
SELECT tr.match_id, tr.team, tr.rounds_won, bp.best_rating
FROM team_rounds tr
LEFT JOIN best_player bp ON bp.match_id = tr.match_id AND bp.team = tr.team;
"""
SQL_ROUNDS_BY_TEAM_PER_DAY = """
WITH t AS (
  SELECT m.match_date::date AS d, m.team_1 AS team, r.result_1 AS rw
  FROM results r JOIN matches m ON m.match_id = r.match_id
  WHERE r.event_id = :event_id
  UNION ALL
  SELECT m.match_date::date, m.team_2, r.result_2
  FROM results r JOIN matches m ON m.match_id = r.match_id
  WHERE r.event_id = :event_id
)
SELECT d, team, SUM(rw) AS rounds_won
FROM t
GROUP BY d, team
ORDER BY d, team;
"""

def pie_chart(event_id=2208):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_PIE, {"event_id": event_id})
    fig, ax = plt.subplots()
    if df.empty: save_plot(df, fig, f"pie_maps_{es}.png", "pie"); return
    ax.pie(df["maps_played"], labels=df["map"], autopct="%1.1f%%")
    ax.set_title(f"Map distribution — {ename}")
    save_plot(df, fig, f"pie_maps_{es}.png", "pie")

def bar_chart(event_id=2335):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_BAR, {"event_id": event_id})
    fig, ax = plt.subplots()
    if df.empty: save_plot(df, fig, f"bar_team_wins_{es}.png", "bar"); return
    ax.bar(df["team"], df["wins"])
    ax.set_title(f"Top teams by match wins — {ename}")
    ax.set_xlabel("Team"); ax.set_ylabel("Match wins")
    ax.tick_params(axis='x', labelrotation=45)
    for lb in ax.get_xticklabels(): lb.set_horizontalalignment('right')
    save_plot(df, fig, f"bar_team_wins_{es}.png", "bar")

def barh_chart(event_id=2335, min_maps=8):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_BARH, {"event_id": event_id, "min_maps": min_maps})
    fig, ax = plt.subplots()
    if df.empty: save_plot(df, fig, f"barh_players_rating_{es}.png", "barh"); return
    labels = df["player_name"] + " (" + df["team"].fillna("—") + ")"
    ax.barh(labels, df["avg_rating"]); ax.invert_yaxis()
    ax.set_title(f"Players by avg rating (≥{min_maps}) — {ename}")
    ax.set_xlabel("Average rating")
    save_plot(df, fig, f"barh_players_rating_{es}.png", "barh")

def line_chart(team="Natus Vincere", year=2019):
    df = fetch_df(SQL_LINE, {"team": team, "year": year})
    fig, ax = plt.subplots()
    if df.empty: save_plot(df, fig, f"line_{slug(team)}_{year}.png", "line"); return
    ax.plot(df["d"], df["rounds_won"], marker="o")
    ax.set_title(f"{team}: rounds won over time in {year}")
    ax.set_xlabel("Date"); ax.set_ylabel("Rounds won")
    save_plot(df, fig, f"line_{slug(team)}_{year}.png", "line")

def hist_chart(event_id=2208, bins=15):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_HIST, {"event_id": event_id})
    fig, ax = plt.subplots()
    if df.empty: save_plot(df, fig, f"hist_total_rounds_{es}.png", "hist"); return
    ax.hist(df["total_rounds"], bins=bins)
    ax.set_title(f"Total rounds per map — {ename}")
    ax.set_xlabel("Total rounds"); ax.set_ylabel("Frequency")
    save_plot(df, fig, f"hist_total_rounds_{es}.png", "hist")

def scatter_chart(event_id=2208):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_SCATTER, {"event_id": event_id}).dropna(subset=["best_rating"])
    fig, ax = plt.subplots()
    if df.empty: save_plot(df, fig, f"scatter_rating_vs_rounds_{es}.png", "scatter"); return
    ax.scatter(df["best_rating"], df["rounds_won"])
    ax.set_title(f"Best player rating vs team rounds — {ename}")
    ax.set_xlabel("Best player rating"); ax.set_ylabel("Team rounds won")
    save_plot(df, fig, f"scatter_rating_vs_rounds_{es}.png", "scatter")

def pxy_line_rounds_by_team(event_id=2208):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_ROUNDS_BY_TEAM_PER_DAY, {"event_id": event_id})
    if df.empty: print("[WARN] no data for line"); return
    teams = sorted(df["team"].dropna().unique().tolist())
    fig = px.line(df, x="d", y="rounds_won", color="team",
                  title=f"Rounds won over time — {ename}",
                  hover_data={"team": True, "d": True, "rounds_won": True})
    fig.update_xaxes(rangeslider=dict(visible=True))
    buttons, all_label = [], "ALL"
    for t in [all_label] + teams:
        visible = [True if t == all_label else (tr.name == t) for tr in fig.data]
        buttons.append(dict(label=t, method="update",
                            args=[{"visible": visible},
                                  {"title": f"Rounds won over time — {ename} [{t}]"}]))
    fig.update_layout(updatemenus=[dict(type="dropdown", x=1.03, y=1, buttons=buttons)])
    save_html(fig, f"plotly_line_rounds_by_team_{es}.html")

def pxy_hist_total_rounds(event_id=2208):
    ename = get_event_name(event_id); es = slug(ename)
    df = fetch_df(SQL_HIST, {"event_id": event_id})
    if df.empty: print("[WARN] no data for hist"); return
    fig = px.histogram(df, x="total_rounds", nbins=20,
                       title=f"Total rounds per map — {ename}")
    fig.update_xaxes(rangeslider=dict(visible=True))
    save_html(fig, f"plotly_hist_total_rounds_{es}.html")

# --- Excel ---
def export_to_excel(sheets: dict, filename: str):
    out_path = os.path.join("exports", filename)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=name, index=False)
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"; ws.auto_filter.ref = ws.dimensions
        if ws.max_row >= 2 and ws.max_column >= 1:
            num_cols = []
            for col in range(1, ws.max_column + 1):
                sample = []
                for r in range(2, min(20, ws.max_row) + 1):
                    v = ws.cell(row=r, column=col).value
                    if isinstance(v, (int, float)) and not isinstance(v, bool): sample.append(v)
                if len(sample) >= 3: num_cols.append(col)
            for col in num_cols:
                c = ws.cell(row=1, column=col).column_letter
                rng = f"{c}2:{c}{ws.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(rng, rule)
    wb.save(out_path)
    total = sum(len(df) for df in sheets.values() if df is not None)
    print(f"[OK] Excel {filename} with {len(sheets)} sheets, {total} rows")



if __name__ == "__main__":

    # PNG
    pie_chart(event_id=2208)
    bar_chart(event_id=2335)
    barh_chart(event_id=2335, min_maps=8)
    line_chart(team="Natus Vincere", year=2019)
    hist_chart(event_id=2208, bins=15)
    scatter_chart(event_id=2208)

    # HTML 
    pxy_line_rounds_by_team(event_id=2208)
    pxy_hist_total_rounds(event_id=2208)

    # Excel
    en1 = slug(get_event_name(2208)); en2 = slug(get_event_name(2335))
    dfs = {
        f"maps_{en1}": fetch_df(SQL_PIE, {"event_id": 2208}),
        f"team_wins_{en2}": fetch_df(SQL_BAR, {"event_id": 2335}),
        f"players_rating_{en2}": fetch_df(SQL_BARH, {"event_id": 2335, "min_maps": 8}),
    }
    export_to_excel(dfs, "csgo_report.xlsx")

